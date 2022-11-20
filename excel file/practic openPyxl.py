import openpyxl
from openpyxl import*
from openpyxl.styles import*
import xlrd
import xlwt
#Task1
sd=load_workbook("Ara Sınıf Erkek Yedek Liste.xlsx")
stdList = sd.worksheets[0]
print(stdList)
name=stdList['F19'].value
print(name)
cell= stdList['A18': 'E30']
print()
for i in cell:
    for j in i:
        print(j.value, end =' ')
    print()
#Task6
list_m=openpyxl.load_workbook('Ara Sınıf Erkek Yedek Liste.xlsx')
ch=list_m.active
for i in range(1,ch.max_row+1):
    print('\n')
    print("Строка",i,"данные:")
    for j in range(1,ch.max_column+1):
        cell_obj=ch.cell(row=i,column=j)
        print(cell_obj.value,end="")
#Task7
list_m=openpyxl.load_workbook('Ara Sınıf Erkek Yedek Liste.xlsx')
ch=list_m.active
for i in range(1,6):
    print('\n')
    print("Строка",i,"данные:")
    for j in range(1,4):
        cell_obj=ch.cell(row=i,column=j)
        print(cell_obj.value,end="")
#Task8
list1=[]
list2=[]
list3=[]
list4=[]
for i in range(2):
    name = input("Введите имя:")
    age=int(input("Введите возрост:"))
    salary=int(input("Введите зарплата:"))
    yearWorks=int(input("Введите Год работ:"))
    list1.append(name)
    list2.append(age)
    list3.append(salary)
    list4.append(yearWorks)

a=print(list1)
b=print(list2)
c=print(list3)
d=print(list4)
Task9
list1=['Самат','Улан','Asan']
list2=[25,30,21]
list3=[25000,35000,45000]
list4=[5,8,9]
workers= Workbook()
sheet2 = workers.worksheets[0]
cell1Header = sheet2.cell(row = 1, column = 1).value = ' Имя'
cell2Header = sheet2.cell(row = 1, column = 2).value = 'Возрост'
cell3Header = sheet2.cell(row = 1, column = 3).value = 'Зарплата'
cell4Header = sheet2.cell(row = 1, column = 4).value = 'Год работ'
sheet2['A1'].font = Font(bold=True)
sheet2['B1'].font = Font(bold=True)
sheet2['C1'].font = Font(bold=True)
sheet2['D1'].font = Font(bold=True)

counter = 3
for i in range(3):
    sheet2.cell(row = counter, column = 1).value = list1[i]
    sheet2.cell(row = counter, column = 2).value = list2[i]
    sheet2.cell(row = counter, column = 3).value = list3[i]
    sheet2.cell(row=counter, column=4).value = list4[i]
    counter += 1
workers.save('workrs2.xlsx')
